"""Build the NLA manual-research template spreadsheet.

Input : data/es/malaga/facilities/malaga_facilities.json (99 facilities)
Output: data/es/malaga/facilities/malaga_nla_research.xlsx

Re-runnable — overwrites the xlsx each run. Safe because the user
fills the researched file in place under a different name
(malaga_nla_research_filled.xlsx) which this script doesn't touch.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
SRC = HERE / "malaga_facilities.json"
OUT = HERE / "malaga_nla_research.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="2EC4A0")  # Veridian green
SUB_FILL = PatternFill("solid", start_color="E8F5F0")
BLANK_FILL = PatternFill("solid", start_color="FFF9E6")  # soft yellow for TO-FILL
INSTR_FILL = PatternFill("solid", start_color="F5F6F2")
NOTE_FILL = PatternFill("solid", start_color="FDECEC")  # soft red for cautions

BORDER = Border(
    left=Side(style="thin", color="D5D7D0"),
    right=Side(style="thin", color="D5D7D0"),
    top=Side(style="thin", color="D5D7D0"),
    bottom=Side(style="thin", color="D5D7D0"),
)

data = json.loads(SRC.read_text(encoding="utf-8"))
facilities = data["facilities"]

wb = Workbook()


# ───────── Sheet 1 : Instructions ─────────
ws1 = wb.active
ws1.title = "Instructions"
ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 110

rows = [
    ("TITLE", "Project Veridian — Málaga NLA Manual Research"),
    ("SUB", "For facilities in data/es/malaga/facilities/malaga_facilities.json (99 sites)"),
    ("BLANK", ""),
    ("H2", "1. What you're filling in"),
    ("P", "The Facilities sheet has one row per self-storage facility in Málaga province. Columns in "
          "yellow are empty and need researching. Everything else (name, address, lat/lng, website, etc.) "
          "is already populated from Google Places and should not be changed."),
    ("BLANK", ""),
    ("H3", "nla_sqm  (column N)"),
    ("P", "Net Lettable Area in m² — the total floor area available to rent out as storage units, "
          "excluding corridors, stairs, lifts, offices, loading bays, plant rooms. If the operator "
          "publishes a figure directly, use it exactly. Otherwise estimate as described in §3."),
    ("H3", "constructed_area_sqm  (column O)"),
    ("P", "Total gross constructed floor area (superficie construida). Preferred source is Catastro (§3). "
          "If you can only find NLA, leave this blank and the pipeline will estimate it as nla_sqm / 0.70. "
          "If you can only find constructed, pipeline estimates NLA as constructed × 0.70."),
    ("H3", "confidence  (column P, dropdown)"),
    ("P", "high = operator publishes the figure on its website or in press releases. "
          "medium = derived from Catastro constructed-area with NLA-ratio assumption. "
          "low = visual estimate from satellite imagery / StreetView / square-footage guess."),
    ("H3", "source_url  (column Q)"),
    ("P", "URL of the primary source you used. If multiple, pick the most authoritative. "
          "Leave blank only if source is Catastro (reference by Referencia Catastral in the notes column)."),
    ("H3", "notes  (column R)"),
    ("P", "Anything the pipeline or future-you should know: Referencia Catastral, number of units, "
          "number of floors, whether the facility is inside a shared building, whether it's a "
          "'guardamuebles' (bulk storage, not self-service) — anything that explains a non-obvious value."),
    ("BLANK", ""),
    ("H2", "2. What NOT to fill in"),
    ("P", "Do NOT edit: id, name, brand, municipio, postal_code, address, lat, lng, place_id, "
          "phone, website, rating, rating_count, source. These are from Google Places and will be "
          "merged back into the JSON unchanged."),
    ("P", "Do NOT try to fill facility_type (self_storage vs guardamuebles) or size_tier — the "
          "downstream pipeline (scripts/prepare-data.js) classifies those automatically from the "
          "NLA value and operator name regex. Your only job is NLA + constructed area."),
    ("BLANK", ""),
    ("H2", "3. Where to find the data (in order of preference)"),
    ("H3", "3a. Operator website — HIGH confidence"),
    ("P", "Chains almost always publish total m² or unit counts per location. Quick-reference URLs "
          "are in the Operator_Sources sheet. Look for pages titled 'Nuestros centros', 'Instalaciones', "
          "'Ubicaciones', 'About', capacity tables, or individual facility 'ficha' pages."),
    ("P", "Brands in Málaga (with facility counts): Trasteros Plus 10, Keepy 6, Islatrans 6, "
          "trasteroZ 3, EasyBox 3, Reva 3, uStore-it 3, Full Storage 3, Dan 2, Trasteros Axarquía 2, "
          "WARDA 2, Independent 56. For 'Independent' facilities each needs its own research."),
    ("H3", "3b. Catastro — MEDIUM confidence (via constructed_area × 0.70)"),
    ("P", "https://www.sedecatastro.gob.es → 'Consulta de datos catastrales' → 'Por localización'. "
          "Enter the address (Málaga province, muni, street, number) or click directly on the map. "
          "The 'Datos descriptivos' panel shows 'Superficie construida' in m². Copy Referencia "
          "Catastral (14 characters) into the notes column for traceability."),
    ("P", "Catastro gives GROSS constructed area. For self-storage the NLA typically sits at "
          "65–80 % of constructed (depends on how many corridors / offices). Use 0.70 as a default "
          "multiplier. Flag this row as confidence=medium."),
    ("H3", "3c. Google Maps / Street View / satellite — LOW confidence"),
    ("P", "Last resort for Independent single-site operators with no web presence and no clean "
          "Catastro match. Measure the building footprint on Google Maps (right-click → 'Medir "
          "distancia'), multiply by visible floor count from StreetView, then by 0.70. Flag the "
          "row confidence=low and explain the estimation in the notes column."),
    ("H3", "3d. When the facility is clearly a 'guardamuebles' (removals / bulk storage)"),
    ("P", "If the business is primarily a mover/guardamuebles (Islatrans, Trasteros Axarquía, some "
          "Independents), NLA as a storage KPI is less meaningful — they rent by volume-week not "
          "m²-month. Still provide the best area estimate you can; the pipeline will tag the "
          "facility as guardamuebles based on operator-name regex and constructed_area_sqm "
          "presence (see scripts/prepare-data.js)."),
    ("BLANK", ""),
    ("H2", "4. How this gets merged back"),
    ("P", "When you're done, save the filled workbook as malaga_nla_research_filled.xlsx (don't "
          "overwrite the template). Claude will then run a merge script that reads the xlsx, maps "
          "facility ids to the corresponding entries in malaga_facilities.json, copies nla_sqm and "
          "constructed_area_sqm (and renames Málaga's nla_m2 field to match Euskadi's schema), "
          "and commits the enriched JSON."),
    ("P", "After that, the downstream pipeline (scripts/prepare-data.js — separate branch) will "
          "aggregate NLA per muni, compute nla_per_capita / nla_per_1000_households, classify "
          "facility_type and size_tier, and feed the opportunity-score formula."),
    ("BLANK", ""),
    ("H2", "5. Sanity bounds (rough)"),
    ("P", "Typical self-storage NLA ranges — use these as smell-tests, not hard limits:"),
    ("P", "    small    50 – 300 m²    (converted garage / bajo comercial; often Independents)"),
    ("P", "    medium   300 – 1,500 m² (typical urban Trasteros Plus, Keepy; 1 floor)"),
    ("P", "    large    1,500 – 5,000 m² (regional uStore-it, multi-floor purpose-built)"),
    ("P", "    xlarge   5,000 m²+       (rare in Málaga — mostly Marbella / outskirts of Málaga city)"),
    ("P", "If your value falls outside this range, double-check the source before committing."),
    ("BLANK", ""),
    ("H2", "6. Contact"),
    ("P", "Project lead: Tomas Ru <tomasrugeroni@gmail.com>. Flag ambiguities in the notes column "
          "rather than guessing — it's cheaper to mark a row confidence=low than to revisit later."),
]

r = 1
for kind, text in rows:
    cell = ws1.cell(row=r, column=2, value=text)
    if kind == "TITLE":
        cell.font = Font(name=FONT, size=16, bold=True, color="2A2D26")
        ws1.row_dimensions[r].height = 28
    elif kind == "SUB":
        cell.font = Font(name=FONT, size=11, italic=True, color="547D74")
    elif kind == "H2":
        cell.font = Font(name=FONT, size=13, bold=True, color="409B7E")
        cell.fill = INSTR_FILL
        ws1.row_dimensions[r].height = 22
    elif kind == "H3":
        cell.font = Font(name=FONT, size=11, bold=True, color="547D74")
    elif kind == "P":
        cell.font = Font(name=FONT, size=10, color="2A2D26")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[r].height = max(18, 16 * (1 + len(text) // 95))
    r += 1


# ───────── Sheet 2 : Facilities ─────────
ws2 = wb.create_sheet("Facilities")

headers = [
    # Locked — from Google Places, DO NOT EDIT
    ("id", 10, False),
    ("name", 36, False),
    ("brand", 18, False),
    ("municipio", 20, False),
    ("postal_code", 10, False),
    ("address", 40, False),
    ("lat", 11, False),
    ("lng", 11, False),
    ("place_id", 28, False),
    ("phone", 16, False),
    ("website", 40, False),
    ("rating", 7, False),
    ("rating_count", 10, False),
    # To fill
    ("nla_sqm", 11, True),
    ("constructed_area_sqm", 13, True),
    ("confidence", 12, True),
    ("source_url", 40, True),
    ("notes", 60, True),
]

for col_idx, (name, width, _) in enumerate(headers, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=name)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[1].height = 32

for row_idx, f in enumerate(facilities, start=2):
    for col_idx, (name, _, to_fill) in enumerate(headers, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=f.get(name))
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=(name in ("address", "name", "notes")))
        cell.border = BORDER
        if to_fill:
            cell.fill = BLANK_FILL
        if name in ("lat", "lng"):
            cell.number_format = "0.0000000"
        if name == "rating":
            cell.number_format = "0.0"
        if name == "nla_sqm" or name == "constructed_area_sqm":
            cell.number_format = "#,##0"
        # Hyperlink the website + source_url columns
        if name == "website" and f.get(name):
            cell.hyperlink = f[name]
            cell.font = Font(name=FONT, size=10, color="0070C0", underline="single")

# Freeze header row + lock first column
ws2.freeze_panes = "B2"

# Turn the facilities range into an Excel table (enables sort/filter)
last_col_letter = get_column_letter(len(headers))
last_row = len(facilities) + 1
table = Table(displayName="Facilities", ref=f"A1:{last_col_letter}{last_row}")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
)
ws2.add_table(table)

# Dropdown on confidence column (col P = 16)
conf_col_letter = get_column_letter(headers.index(("confidence", 12, True)) + 1)
dv = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
dv.error = "Pick one of: high, medium, low"
dv.errorTitle = "Invalid confidence"
dv.add(f"{conf_col_letter}2:{conf_col_letter}{last_row}")
ws2.add_data_validation(dv)


# ───────── Sheet 3 : Operator_Sources ─────────
ws3 = wb.create_sheet("Operator_Sources")
ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 7
ws3.column_dimensions["C"].width = 52
ws3.column_dimensions["D"].width = 55

# Count facilities per brand
from collections import Counter

brand_counts = Counter(f["brand"] for f in facilities)

# Headers
for col_idx, h in enumerate(["Brand", "# sites", "Website(s) to check first", "Notes"], start=1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
ws3.row_dimensions[1].height = 22

# Known operator URLs (seed list — researcher can add more)
operator_urls = {
    "Trasteros Plus": ("https://www.trasterosplus.es/",
                       "Has per-location pages; look under 'Nuestros centros' or each ficha for m² / unit counts."),
    "Keepy": ("https://www.keepy.es/",
              "Facility pages usually list total m² and number of units per size band."),
    "Islatrans": ("https://www.islatrans.com/",
                  "Primarily moving/guardamuebles — area may be published as volume (m³) rather than m²."),
    "trasteroZ": ("https://trasteroz.com/",
                  "Small chain; each location has its own page."),
    "EasyBox": ("https://www.easybox.es/",
                "Check 'Centros' list."),
    "Reva": ("https://reva.es/",
             "Smaller regional operator."),
    "uStore-it": ("https://www.ustore-it.com/",
                  "Largest purpose-built facilities in the sample — typically multi-floor, 2,000+ m²."),
    "Full Storage": ("https://www.fullstorage.es/",
                     "Check 'Ubicaciones'."),
    "Dan": ("",  # unknown
            "Search for 'Dan Storage' or 'Dan Trastero' + Málaga; 2 sites — may be independent."),
    "Trasteros Axarquía": ("",
                           "Regional chain covering the Axarquía (Vélez-Málaga area) — likely guardamuebles."),
    "WARDA": ("",
              "2 facilities — check if single-owner local operator."),
    "Independent": ("",
                    "56 facilities — each needs its own research. Use Catastro for these."),
}

r = 2
for brand, cnt in brand_counts.most_common():
    url, note = operator_urls.get(brand, ("", ""))
    ws3.cell(row=r, column=1, value=brand).font = Font(name=FONT, size=10, bold=(brand != "Independent"))
    ws3.cell(row=r, column=2, value=cnt).alignment = Alignment(horizontal="center")
    c_url = ws3.cell(row=r, column=3, value=url)
    if url:
        c_url.hyperlink = url
        c_url.font = Font(name=FONT, size=10, color="0070C0", underline="single")
    else:
        c_url.font = Font(name=FONT, size=10, italic=True, color="8E8E8E")
    note_cell = ws3.cell(row=r, column=4, value=note)
    note_cell.font = Font(name=FONT, size=10)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 5):
        ws3.cell(row=r, column=col).border = BORDER
    ws3.row_dimensions[r].height = 28
    r += 1

# Catastro row as a separator + reference block
r += 1
ws3.cell(row=r, column=1, value="Catastro (fallback)").font = Font(name=FONT, size=10, bold=True, color="409B7E")
ws3.cell(row=r, column=3, value="https://www.sedecatastro.gob.es/").hyperlink = (
    "https://www.sedecatastro.gob.es/"
)
ws3.cell(row=r, column=3).font = Font(name=FONT, size=10, color="0070C0", underline="single")
ws3.cell(row=r, column=4, value="Consulta por localización → superficie construida. Copy Referencia Catastral into notes.").font = Font(name=FONT, size=10)
ws3.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
ws3.row_dimensions[r].height = 28


# ───────── Save ─────────
wb.save(OUT)
print(f"Wrote {OUT.relative_to(HERE.parent.parent.parent)} ({OUT.stat().st_size // 1024} KB)")
print(f"  Sheets: Instructions, Facilities ({len(facilities)} rows), Operator_Sources")
print(f"  Editable columns (yellow): nla_sqm, constructed_area_sqm, confidence, source_url, notes")
